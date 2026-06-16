import base64
import csv
import difflib
import io
import os
import re
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from functools import lru_cache
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels
from PIL import Image as PILImage
from PIL import ImageOps, UnidentifiedImageError
from pypdf import PdfReader


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_CSV = BASE_DIR / "data/Listagem dos produtos.xlsx.csv"
DEFAULT_ZIP = BASE_DIR / "data/Fotos Cod-20260430T175121Z-3-001.zip"
DEFAULT_PHOTOS = BASE_DIR / "Fotos Cod"
DEFAULT_FACTORY_CODES = BASE_DIR / "produtos codigo fabrica.xlsx"
IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp")
TEMPLATE_SUFFIX = "FICHA CADASTRO PRODUTO C&M.xlsx"
PRICE_FILE = BASE_DIR / "preço.xlsx"
DEFAULT_PRICE_REGION = 1
MATCH_STOP_TOKENS = {
    "FICHA",
    "FICHAS",
    "FTC",
    "TECNICA",
    "TECNICAS",
    "PRODUTO",
    "PRODUTOS",
    "CARNE",
    "CARNES",
    "CONG",
    "CONGELADO",
    "CONGELADA",
    "RESF",
    "RESFRIADO",
    "RESFRIADA",
    "BOVINO",
    "BOVINA",
    "BOV",
    "SUINO",
    "SUINA",
    "CP",
    "CG",
    "RF",
    "VC",
    "TIMBRADA",
    "CARTUCHO",
    "CAIXA",
    "CAIXAS",
    "FRIGORIFICO",
    "LTDA",
    "S/A",
    "COM",
    "SEM",
    "PARA",
    "DAS",
    "DOS",
    "KG",
    "UN",
    "UND",
    "COZIDO",
    "COZIDA",
    "COZIDOS",
    "COZIDAS",
    "TEMPERADO",
    "TEMPERADA",
    "TEMPERADOS",
    "TEMPERADAS",
    "DESF",
    "DESFIADO",
    "DESFIADA",
    "DESFIADOS",
    "DESFIADAS",
}

PRODUCT_IDENTITY_TOKENS = {
    "ACEM",
    "ALCATRA",
    "ANCHO",
    "BARRIGA",
    "BOLINHO",
    "BOMBOM",
    "BRIE",
    "BROCOLIS",
    "CALABRESA",
    "CHORIZO",
    "CONTRA",
    "COSTELA",
    "COXA",
    "COXAO",
    "COXINHA",
    "CUPIM",
    "DENVER",
    "ENTRECOTE",
    "FILE",
    "FILEZINHO",
    "FRALDA",
    "FRALDAO",
    "FRANGO",
    "GOUDA",
    "HAMBURGUER",
    "LOMBO",
    "MAMINHA",
    "MANDIOCA",
    "MIGNON",
    "MORANGO",
    "MUSCULO",
    "PALETA",
    "PANCETA",
    "PATINHO",
    "PEITO",
    "PERNIL",
    "PICANHA",
    "REQUEIJAO",
    "SOBREPALETA",
    "SOL",
    "VAGEM",
}


def find_template_file():
    return next(
        (p for p in BASE_DIR.glob("*.xlsx") if p.name.endswith(TEMPLATE_SUFFIX)),
        BASE_DIR / TEMPLATE_SUFFIX,
    )


def unique_headers(headers):
    seen = {}
    out = []
    for h in headers:
        base = (h or "").strip() or "Coluna"
        seen[base] = seen.get(base, 0) + 1
        out.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return out


def normalize_key(value):
    value = unicodedata.normalize("NFKD", str(value or ""))
    value = value.encode("ascii", "ignore").decode("ascii")
    value = value.lower()
    return re.sub(r"[^a-z0-9]+", "", value)


KNOWN_PRODUCT_BRANDS = [
    "CARAPRETA",
    "ALFAMA",
    "FRIELLA",
    "COOPAVEL",
    "VALENCIO",
    "ATIGEL",
    "AVE NOVA",
    "GUIDARA",
    "DUBOI",
    "DAUS",
    "EASYCHEF",
    "MARIZA",
    "MINERVA",
    "FRIBOI",
    "MOCOCA",
    "PAMPLONA",
    "PLENA",
    "TUDBOM",
    "SADIA",
    "PERDIGAO",
    "PERDIGÃO",
    "AURORA",
    "RIO MARIA",
    "RAINHA DA PAZ",
    "SOMAVE",
    "STELLADORO",
    "SAO FRANCISCO",
    "SÃO FRANCISCO",
    "SAO VICENTE",
    "SÃO VICENTE",
    "TEMPERO DA CASA",
    "BRF",
]


def normalize_brand_text(value):
    value = unicodedata.normalize("NFKD", str(value or ""))
    value = value.encode("ascii", "ignore").decode("ascii")
    value = value.upper()
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def detect_known_brand(*values):
    text = normalize_brand_text(" ".join(str(value or "") for value in values))
    if not text:
        return ""

    compact_text = normalize_key(text)
    for brand in KNOWN_PRODUCT_BRANDS:
        normalized_brand = normalize_brand_text(brand)
        compact_brand = normalize_key(brand)
        if re.search(rf"\b{re.escape(normalized_brand)}\b", text):
            return brand
        if len(compact_brand) >= 6 and compact_brand in compact_text:
            return brand

    return ""


def read_csv_bytes(data: bytes):
    if not data:
        raise ValueError("O CSV esta vazio.")

    text = None
    for enc in ("utf-8-sig", "cp1252", "latin1"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            pass

    if text is None:
        raise ValueError("Nao consegui ler a codificacao do CSV.")

    sample = text[:5000]
    if not sample.strip():
        raise ValueError("O CSV nao possui conteudo para importar.")

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t,")
        reader = csv.reader(io.StringIO(text), dialect)
    except csv.Error:
        reader = csv.reader(io.StringIO(text), delimiter=";")

    try:
        raw_headers = next(reader)
    except StopIteration as exc:
        raise ValueError("O CSV nao possui linha de cabecalho.") from exc

    headers = unique_headers(raw_headers)
    rows = []
    for row in reader:
        if not any(str(c).strip() for c in row):
            continue
        row = row + [""] * (len(headers) - len(row))
        rows.append({
            headers[i]: row[i] if i < len(row) else ""
            for i in range(len(headers))
        })

    return headers, rows


def read_xlsx_products(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Produtos"] if "Produtos" in wb.sheetnames else wb[wb.sheetnames[0]]

    raw_headers = [
        cell.value if cell.value is not None else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    headers = unique_headers([str(h).strip() for h in raw_headers])

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = ["" if value is None else value for value in row]
        if not any(str(value).strip() for value in values):
            continue
        values = values + [""] * (len(headers) - len(values))
        rows.append({
            headers[i]: values[i] if i < len(values) else ""
            for i in range(len(headers))
        })

    return headers, rows


def read_factory_code_map(path=DEFAULT_FACTORY_CODES):
    path = Path(path)
    if not path.exists():
        return {}

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["produtos codigo fabrica"] if "produtos codigo fabrica" in wb.sheetnames else wb.active
    headers = [
        normalize_key(cell.value)
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    try:
        internal_index = headers.index("codigointerno")
        factory_index = headers.index("codigofabrica")
    except ValueError:
        wb.close()
        return {}

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        internal_code = normalize_code(row[internal_index] if internal_index < len(row) else "")
        factory_code = normalize_factory_code(row[factory_index] if factory_index < len(row) else "")
        if internal_code and factory_code:
            mapping[internal_code] = factory_code

    wb.close()
    return mapping


def apply_factory_codes(rows):
    factory_codes = read_factory_code_map()
    if not factory_codes:
        return rows

    for row in rows:
        code = row_code(row)
        if code and code in factory_codes:
            row.setdefault("Codigo fabrica", factory_codes[code])

    return rows


def normalize_code(value):
    if value is None:
        return ""

    value = str(value).strip()
    if not value:
        return ""

    value = re.sub(r"\.0$", "", value)
    digits = re.findall(r"\d+", value)
    return digits[0].lstrip("0") if digits else value


def normalize_factory_code(value):
    if value is None:
        return ""

    value = str(value).strip().upper()
    if not value:
        return ""

    value = re.sub(r"\.0$", "", value)
    return re.sub(r"[^A-Z0-9]+", "", normalize_brand_text(value))


def format_measure_cm(value_mm):
    value_cm = float(value_mm) / 10
    if value_cm.is_integer():
        return str(int(value_cm))
    return f"{value_cm:.1f}".replace(".", ",")


def format_measure_value(value):
    value = float(value)
    if value.is_integer():
        return str(int(value))
    return f"{value:.1f}".replace(".", ",")


def parse_dimension_number(value):
    raw = str(value or "").strip().replace(",", ".")
    if re.fullmatch(r"\d{4,}", raw) and raw.endswith("50"):
        fixed = raw[:-2]
        if fixed:
            return float(fixed)
    return float(raw)


def parse_box_dimensions(value):
    text = str(value or "")
    labeled = {}
    for label, number, unit in re.findall(
        r"\b(comprimento|largura|altura)\b\s*:?\s*"
        r"(\d+(?:[,.]\d+)?)\s*(mm|cm)\b",
        text,
        flags=re.IGNORECASE,
    ):
        labeled[normalize_key(label)] = (
            format_measure_cm(float(number.replace(",", ".")))
            if unit.lower() == "mm"
            else format_measure_value(number.replace(",", "."))
        )
    if all(key in labeled for key in ("comprimento", "largura", "altura")):
        return {
            "comprimento": labeled["comprimento"],
            "largura": labeled["largura"],
            "altura": labeled["altura"],
        }

    matches = re.findall(
        r"(\d+(?:[,.]\d+)?)\s*(mm|cm)\b",
        text,
        flags=re.IGNORECASE,
    )
    if len(matches) < 3:
        compact_match = re.search(
            r"\b(\d+(?:[,.]\d+)?)\s*(mm|cm)?\s*[xX]\s*"
            r"(\d+(?:[,.]\d+)?)\s*(mm|cm)?\s*[xX]\s*"
            r"(\d+(?:[,.]\d+)?)\s*(mm|cm)?\b",
            text,
            flags=re.IGNORECASE,
        )
        if compact_match:
            groups = compact_match.groups()
            values = [float(groups[i].replace(",", ".")) for i in (0, 2, 4)]
            units = [groups[i].lower() for i in (1, 3, 5) if groups[i]]
            unit = units[-1] if units else ("cm" if max(values) <= 100 else "mm")
            matches = [(str(v).replace(".", ","), unit) for v in values]
    if len(matches) < 3:
        pair_matches = re.findall(
            r"\b(\d+(?:[,.]\d+)?)\s*(mm|cm)?\s*[xX]\s*"
            r"(\d+(?:[,.]\d+)?)\s*(mm|cm)?\b",
            text,
            flags=re.IGNORECASE,
        )
        if pair_matches:
            candidates = []
            for first, first_unit, second, second_unit in pair_matches:
                values = [parse_dimension_number(first), parse_dimension_number(second)]
                unit = (second_unit or first_unit or ("cm" if max(values) <= 100 else "mm")).lower()
                comparable = [value / 10 if unit == "mm" else value for value in values]
                if min(comparable) < 5 or max(comparable) > 200:
                    continue
                converted = [format_measure_value(value) for value in comparable]
                candidates.append((comparable[0] * comparable[1], converted))

            if candidates:
                _, converted = sorted(candidates, key=lambda item: item[0], reverse=True)[0]
                return {
                    "largura": converted[0],
                    "comprimento": converted[1],
                }

        return {}

    converted = []
    for number, unit in matches[:3]:
        number = float(number.replace(",", "."))
        converted.append(
            format_measure_cm(number) if unit.lower() == "mm" else format_measure_value(number)
        )

    return {
        "comprimento": converted[0],
        "largura": converted[1],
        "altura": converted[2],
    }


def find_box_dimensions_text(text):
    text = str(text or "")

    dimensions = re.search(
        r"Dimens\w*\s+da\s+caixa\s*:\s*([^\n\r]+)",
        text,
        flags=re.IGNORECASE,
    )
    if dimensions:
        return re.sub(r"\s+", " ", dimensions.group(1)).strip()

    def normalized_line(value):
        value = unicodedata.normalize("NFKD", str(value or ""))
        value = value.encode("ascii", "ignore").decode("ascii")
        return value.upper()

    lines = text.splitlines()
    labeled_candidates = []
    for index, line in enumerate(lines):
        normalized = normalized_line(line)
        if "DIMENS" not in normalized:
            continue
        if not all(word in normalized for word in ("COMPRIMENTO", "LARGURA", "ALTURA")):
            continue

        context = normalized_line(" ".join(lines[max(0, index - 12):index + 4]))
        score = 0
        if "EMBALAGEM SECUNDARIA" in context:
            score += 20
        if "CAIXA" in context or "PAPEL" in context:
            score += 10
        if "EMBALAGEM PRIMARIA" in context and "EMBALAGEM SECUNDARIA" not in context:
            score -= 10
        labeled_candidates.append((score, index, line))

    if labeled_candidates:
        labeled_candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
        line = labeled_candidates[0][2]
        line = re.sub(r"^\s*Dimens\w*\s*:?\s*", "", line, flags=re.IGNORECASE)
        return re.sub(r"\s+", " ", line).strip()

    labeled_patterns = [
        r"DIMENSIONAL\s+DA.*CAIXA",
        r"DIMENS\w*\s+DA.*CAIXA",
        r"DIMENS\w*\s+DA.*EMBALAG",
        r"DIMENSIONES\s+DEL.*EMBALAJE",
        r"MEDIDAS?\s+DA.*CAIXA",
        r"MEDIDAS?\s+DA.*EMBALAG",
    ]
    for index, line in enumerate(lines):
        label_window = " ".join(lines[index:index + 3])
        normalized = unicodedata.normalize("NFKD", label_window).encode("ascii", "ignore").decode("ascii")
        if not any(re.search(pattern, normalized, flags=re.IGNORECASE) for pattern in labeled_patterns):
            continue

        window = " ".join(lines[index:index + 6])
        matches = re.findall(
            r"\b\d+(?:[,.]\d+)?\s*(?:mm|cm)?\s*[xX]\s*"
            r"\d+(?:[,.]\d+)?\s*(?:mm|cm)?"
            r"(?:\s*[xX]\s*\d+(?:[,.]\d+)?\s*(?:mm|cm)?)?\b",
            window,
            flags=re.IGNORECASE,
        )
        if matches:
            return re.sub(r"\s+", " ", " ".join(matches)).strip()

    loose = re.search(
        r"\b\d+(?:[,.]\d+)?\s*mm\s*[xX]\s*"
        r"\d+(?:[,.]\d+)?\s*mm\s*[xX]\s*"
        r"\d+(?:[,.]\d+)?\s*mm\b",
        text,
        flags=re.IGNORECASE,
    )
    return re.sub(r"\s+", " ", loose.group(0)).strip() if loose else ""


def extract_supplier_pdf_data(pdf_bytes):
    if not pdf_bytes:
        return {}

    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        normal_text = "\n".join(page.extract_text() or "" for page in reader.pages)
        layout_text = "\n".join(
            page.extract_text(extraction_mode="layout") or ""
            for page in reader.pages
        )
        text = f"{normal_text}\n{layout_text}"
    except Exception as exc:
        raise ValueError("Nao consegui ler a ficha PDF do fornecedor.") from exc

    data = {}

    ean = re.search(r"\bEAN\s*-?\s*13\s*:\s*(\d{13})\b", text, flags=re.IGNORECASE)
    if not ean:
        ean = re.search(r"\b(\d{13})\b", text)
    if ean:
        data["ean"] = ean.group(1)

    value = find_box_dimensions_text(text)
    if value:
        data["box_dimensions"] = re.sub(r"\s+", " ", value).strip()
        data.update(parse_box_dimensions(value))

    if not data.get("ean") and not data.get("box_dimensions"):
        raise ValueError("Nao encontrei EAN-13 nem dimensoes da caixa nesta ficha PDF.")

    return data


def apply_supplier_pdf_data(row, pdf_data):
    if not pdf_data:
        return row

    row = dict(row)
    if pdf_data.get("ean"):
        row["GTIN Unid.Venda"] = pdf_data["ean"]
        row["EAN Unid. Tributavel"] = pdf_data["ean"]
    if pdf_data.get("box_dimensions"):
        row["Dimensoes da caixa"] = pdf_data["box_dimensions"]
    if pdf_data.get("altura"):
        row["Unid.Altura(cm)"] = pdf_data["altura"]
    if pdf_data.get("largura"):
        row["Unid.Largura(cm)"] = pdf_data["largura"]
    if pdf_data.get("comprimento"):
        row["Unid.Comprim(cm)"] = pdf_data["comprimento"]
    return row


def product_sheet_names(wb):
    return [name for name in wb.sheetnames if name != "Produtos"]


def sheet_product_identity(sheet):
    code = sheet["D32"].value or sheet["D31"].value or ""
    ean = sheet["F32"].value or sheet["F31"].value or ""
    description = sheet["B15"].value or sheet["A32"].value or sheet["A31"].value or ""
    return {
        "code": normalize_code(code),
        "ean": re.sub(r"\D", "", str(ean or "")),
        "description": str(description or "").strip(),
    }


def xml_text(element, name):
    child = element.find(f"{{*}}{name}")
    return str(child.text or "").strip() if child is not None else ""


def read_nfe_xml_items(xml_files):
    items = []
    for xml_file in xml_files or []:
        file_name = xml_file.get("name", "nota.xml")
        try:
            root = ET.fromstring(xml_file.get("data", b""))
        except (ET.ParseError, TypeError) as exc:
            raise ValueError(f"XML inválido: {file_name}") from exc

        for detail in root.findall(".//{*}det"):
            product = detail.find("{*}prod")
            if product is None:
                continue
            cest = re.sub(r"\D", "", xml_text(product, "CEST"))
            if not cest:
                continue
            ean = xml_text(product, "cEAN")
            if normalize_key(ean) in {"semgtin", ""}:
                ean = xml_text(product, "cEANTrib")
            if normalize_key(ean) == "semgtin":
                ean = ""
            items.append({
                "xmlName": file_name,
                "itemNumber": str(detail.attrib.get("nItem", "")),
                "supplierCode": xml_text(product, "cProd"),
                "ean": re.sub(r"\D", "", ean),
                "description": xml_text(product, "xProd"),
                "ncm": re.sub(r"\D", "", xml_text(product, "NCM")),
                "cest": cest,
            })
    if not items:
        raise ValueError("Nenhum CEST foi encontrado nos XMLs enviados.")
    return items


def xml_description_score(sheet_description, xml_description):
    sheet_text = normalize_brand_text(sheet_description)
    xml_text_value = normalize_brand_text(xml_description)
    if not sheet_text or not xml_text_value:
        return 0
    sheet_tokens = set(sheet_text.split()) - MATCH_STOP_TOKENS
    xml_tokens = set(xml_text_value.split()) - MATCH_STOP_TOKENS
    overlap = len(sheet_tokens & xml_tokens) / max(1, min(len(sheet_tokens), len(xml_tokens)))
    sequence = difflib.SequenceMatcher(None, sheet_text, xml_text_value).ratio()
    return round(max(overlap, sequence) * 100)


def match_xml_item(identity, xml_items, used_indexes):
    if identity["ean"]:
        for index, item in enumerate(xml_items):
            if index not in used_indexes and item["ean"] == identity["ean"]:
                return index, "EAN", 100

    if identity["code"]:
        for index, item in enumerate(xml_items):
            if index in used_indexes or normalize_code(item["supplierCode"]) != identity["code"]:
                continue
            description_score = xml_description_score(identity["description"], item["description"])
            if description_score >= 45:
                return index, "Código + descrição", max(95, description_score)

    best_index = None
    best_score = 0
    for index, item in enumerate(xml_items):
        if index in used_indexes:
            continue
        score = xml_description_score(identity["description"], item["description"])
        if score > best_score:
            best_index = index
            best_score = score
    if best_index is not None and best_score >= 45:
        return best_index, "Descrição", best_score
    return None, "", best_score


def nfe_xml_cest_preview(workbook_bytes, xml_files):
    if not workbook_bytes:
        raise ValueError("Envie o Excel gerado para conferir os XMLs.")
    try:
        wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Não consegui ler o Excel enviado.") from exc

    xml_items = read_nfe_xml_items(xml_files)
    used_indexes = set()
    items = []
    for index, sheet_name in enumerate(product_sheet_names(wb)):
        identity = sheet_product_identity(wb[sheet_name])
        xml_index, method, score = match_xml_item(identity, xml_items, used_indexes)
        xml_item = xml_items[xml_index] if xml_index is not None else {}
        if xml_index is not None:
            used_indexes.add(xml_index)
        items.append({
            "index": index + 1,
            "sheetName": sheet_name,
            "productCode": identity["code"],
            "productDescription": identity["description"],
            "productEan": identity["ean"],
            "selected": xml_index is not None,
            "matchMethod": method,
            "score": score,
            "xmlName": xml_item.get("xmlName", ""),
            "xmlItem": xml_item.get("itemNumber", ""),
            "xmlProductCode": xml_item.get("supplierCode", ""),
            "xmlDescription": xml_item.get("description", ""),
            "ean": xml_item.get("ean", ""),
            "ncm": xml_item.get("ncm", ""),
            "cest": xml_item.get("cest", ""),
        })
    wb.close()
    matched_count = sum(1 for item in items if item["selected"])
    return {
        "sheetCount": len(items),
        "xmlCount": len(xml_files or []),
        "xmlItemCount": len(xml_items),
        "matchedCount": matched_count,
        "missingCestCount": len(items) - matched_count,
        "extraXmlItemCount": max(0, len(xml_items) - matched_count),
        "items": items,
    }


def fill_workbook_with_nfe_cest(workbook_bytes, xml_files, selected_indexes=None):
    preview = nfe_xml_cest_preview(workbook_bytes, xml_files)
    try:
        wb = load_workbook(io.BytesIO(workbook_bytes))
    except Exception as exc:
        raise ValueError("Não consegui ler o Excel enviado.") from exc

    use_explicit_selection = selected_indexes is not None
    selected_indexes = set(selected_indexes or [])
    filled = 0
    for item in preview["items"]:
        if not item["selected"] or not item["cest"]:
            continue
        item_index = item["index"] - 1
        if use_explicit_selection and item_index not in selected_indexes:
            continue
        wb[item["sheetName"]]["B19"] = item["cest"]
        filled += 1
    if not filled:
        raise ValueError("Nenhuma ficha foi selecionada para receber o CEST.")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def pdf_data_from_file(pdf_file):
    if not pdf_file:
        return {}, "Sem ficha selecionada."

    try:
        return extract_supplier_pdf_data(pdf_file.get("data", b"")), ""
    except ValueError as exc:
        return {}, str(exc)


def ensure_header_column(ws, header):
    for cell in ws[1]:
        if normalize_key(cell.value) == normalize_key(header):
            return cell.column

    column = ws.max_column + 1
    ws.cell(1, column, header)
    return column


def apply_pdf_data_to_product_sheet(ws, pdf_data):
    if pdf_data.get("ean"):
        ws["F31"] = pdf_data["ean"]
    if pdf_data.get("altura"):
        ws["B25"] = pdf_data["altura"]
    if pdf_data.get("largura"):
        ws["B26"] = pdf_data["largura"]
    if pdf_data.get("comprimento"):
        ws["B27"] = pdf_data["comprimento"]


def ordered_pdf_preview(workbook_bytes, pdf_files):
    if not workbook_bytes:
        raise ValueError("Envie o Excel gerado para conferir as fichas.")

    try:
        wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Nao consegui ler o Excel enviado.") from exc

    sheets = product_sheet_names(wb)
    items = []
    for index, pdf_file in enumerate(pdf_files or []):
        file_name = pdf_file.get("name", f"PDF {index + 1}") if pdf_file else ""
        data, error = pdf_data_from_file(pdf_file)

        sheet_name = sheets[index] if index < len(sheets) else ""
        sheet = wb[sheet_name] if sheet_name else None
        product_description = str(sheet["B15"].value or "") if sheet else ""
        items.append({
            "index": index + 1,
            "pdfName": file_name,
            "sheetName": sheet_name,
            "productCode": str(sheet["D31"].value or "") if sheet else "",
            "productDescription": product_description,
            "brand": detect_known_brand(product_description, sheet_name, file_name),
            "ean": data.get("ean", ""),
            "boxDimensions": data.get("box_dimensions", ""),
            "altura": data.get("altura", ""),
            "largura": data.get("largura", ""),
            "comprimento": data.get("comprimento", ""),
            "error": error,
        })

    wb.close()
    return {
        "sheetCount": len(sheets),
        "pdfCount": len(pdf_files or []),
        "matchedCount": min(len(sheets), len(pdf_files or [])),
        "extraPdfCount": max(0, len(pdf_files or []) - len(sheets)),
        "missingPdfCount": max(0, len(sheets) - len(pdf_files or [])),
        "items": items,
    }


def pdf_file_dict_from_path(path, base_folder=None):
    path = Path(path)
    name = str(path.relative_to(base_folder)) if base_folder else path.name
    return {"name": name, "data": path.read_bytes(), "path": str(path)}


def read_pdf_files_from_folder_selection(pdf_folder_path, selected_files):
    folder = Path(pdf_folder_path).expanduser()
    if not str(pdf_folder_path or "").strip():
        raise ValueError("Informe a pasta das fichas.")
    if not folder.exists():
        raise ValueError(f"Pasta de fichas nao encontrada: {folder}")
    if not folder.is_dir():
        raise ValueError(f"O caminho das fichas nao e uma pasta: {folder}")

    folder = folder.resolve()
    out = []
    for rel_name in selected_files or []:
        if not str(rel_name or "").strip():
            out.append(None)
            continue

        pdf_path = (folder / rel_name).resolve()
        if folder not in pdf_path.parents and pdf_path != folder:
            raise ValueError(f"Ficha fora da pasta informada: {rel_name}")
        if not pdf_path.exists() or not pdf_path.is_file():
            raise ValueError(f"Ficha nao encontrada: {rel_name}")
        if pdf_path.suffix.lower() != ".pdf":
            raise ValueError(f"Arquivo selecionado nao e PDF: {rel_name}")
        out.append(pdf_file_dict_from_path(pdf_path, folder))

    return out


def inspect_pdf_from_folder(pdf_folder_path, selected_file):
    pdf_files = read_pdf_files_from_folder_selection(pdf_folder_path, [selected_file])
    pdf_file = pdf_files[0] if pdf_files else None
    data, error = pdf_data_from_file(pdf_file)
    file_name = pdf_file.get("name", "") if pdf_file else str(selected_file or "")
    return {
        "pdfName": file_name,
        "suggestedFile": file_name,
        "selected": bool(file_name and not error),
        "brand": detect_known_brand(file_name),
        "ean": data.get("ean", ""),
        "boxDimensions": data.get("box_dimensions", ""),
        "altura": data.get("altura", ""),
        "largura": data.get("largura", ""),
        "comprimento": data.get("comprimento", ""),
        "error": error,
    }


def ordered_pdf_folder_suggestions(workbook_bytes, pdf_folder_path, min_score=62):
    if not workbook_bytes:
        raise ValueError("Envie o Excel gerado para sugerir fichas.")

    folder = Path(pdf_folder_path).expanduser()
    if not str(pdf_folder_path or "").strip():
        raise ValueError("Informe a pasta das fichas.")
    if not folder.exists():
        raise ValueError(f"Pasta de fichas nao encontrada: {folder}")
    if not folder.is_dir():
        raise ValueError(f"O caminho das fichas nao e uma pasta: {folder}")

    try:
        wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Nao consegui ler o Excel enviado.") from exc

    folder = folder.resolve()
    pdfs = sorted(folder.rglob("*.pdf"))
    pdf_labels = [
        {
            "path": pdf,
            "file": str(pdf.relative_to(folder)),
            "label": supplier_pdf_text_label(pdf),
        }
        for pdf in pdfs
    ]
    for pdf in pdf_labels:
        pdf["brand"] = supplier_pdf_brand(pdf["path"], pdf["label"])
        pdf["internalCode"] = supplier_pdf_internal_code(pdf["path"])
    factory_codes = read_factory_code_map()
    sheets = product_sheet_names(wb)
    used_files = set()
    items = []

    for index, sheet_name in enumerate(sheets):
        sheet = wb[sheet_name]
        product_code = str(sheet["D31"].value or "")
        product_description = str(sheet["B15"].value or "")
        product_family = str(sheet["A31"].value or "")
        product_label = product_sheet_match_label(sheet_name, product_description, product_family)
        product_brand = detect_known_brand(product_description, product_family, sheet_name)
        product_factory_code = factory_codes.get(normalize_code(product_code), "")

        best = None
        best_score = 0
        for pdf in pdf_labels:
            if pdf["file"] in used_files:
                continue
            score = score_ordered_pdf_suggestion(
                product_label,
                pdf["label"],
                product_brand,
                pdf["brand"],
                product_factory_code,
                product_code,
                pdf["internalCode"],
            )
            if score > best_score:
                best_score = score
                best = pdf

        selected = bool(best and best_score >= min_score)
        data = {}
        error = ""
        if best:
            if selected:
                used_files.add(best["file"])
            data, error = pdf_data_from_file(pdf_file_dict_from_path(best["path"], folder))

        items.append({
            "index": index + 1,
            "sheetName": sheet_name,
            "productCode": product_code,
            "productDescription": product_description,
            "selected": selected,
            "score": best_score,
            "suggestedFile": best["file"] if best else "",
            "pdfName": best["file"] if best else "",
            "brand": detect_known_brand(product_description, sheet_name, best["file"] if best else ""),
            "ean": data.get("ean", ""),
            "boxDimensions": data.get("box_dimensions", ""),
            "altura": data.get("altura", ""),
            "largura": data.get("largura", ""),
            "comprimento": data.get("comprimento", ""),
            "error": error,
        })

    wb.close()
    selected_count = sum(1 for item in items if item["selected"])
    return {
        "folderPath": str(folder),
        "sheetCount": len(sheets),
        "pdfCount": len(pdfs),
        "matchedCount": selected_count,
        "extraPdfCount": max(0, len(pdfs) - selected_count),
        "missingPdfCount": len(sheets) - selected_count,
        "pdfOptions": [pdf["file"] for pdf in pdf_labels],
        "items": items,
    }


def audit_product_pdf_suggestions(pdf_folder_path, folder_path="", min_score=62):
    folder = Path(pdf_folder_path).expanduser()
    if not str(pdf_folder_path or "").strip():
        raise ValueError("Informe a pasta das fichas.")
    if not folder.exists():
        raise ValueError(f"Pasta de fichas nao encontrada: {folder}")
    if not folder.is_dir():
        raise ValueError(f"O caminho das fichas nao e uma pasta: {folder}")

    rows, zip_file, zip_mapping, folder_mapping = load_inputs(folder_path=folder_path)
    try:
        image_codes = set(zip_mapping) | set(folder_mapping)
        products = [row_summary(row, image_codes) for row in rows if row_code(row) in image_codes]
    finally:
        if zip_file:
            zip_file.close()

    folder = folder.resolve()
    pdfs = sorted(folder.rglob("*.pdf"))
    pdf_labels = []
    for pdf in pdfs:
        label = supplier_pdf_text_label(pdf)
        pdf_labels.append({
            "file": str(pdf.relative_to(folder)),
            "label": label,
            "brand": supplier_pdf_brand(pdf, label),
            "internalCode": supplier_pdf_internal_code(pdf),
        })
    used_files = set()
    items = []

    for product in products:
        product_label = product_sheet_match_label(
            product["description"],
            product["supplier"],
            product["brand"],
            product["category"],
        )
        product_brand = detect_known_brand(
            product["description"], product["supplier"], product["brand"]
        )
        best = None
        best_score = 0
        for pdf in pdf_labels:
            if pdf["file"] in used_files:
                continue
            score = score_ordered_pdf_suggestion(
                product_label,
                pdf["label"],
                product_brand,
                pdf["brand"],
                product["factoryCode"],
                product["code"],
                pdf["internalCode"],
            )
            if score > best_score:
                best = pdf
                best_score = score

        suggested_file = best["file"] if best and best_score > 0 else ""
        selected = bool(suggested_file and best_score >= min_score)
        if selected:
            used_files.add(suggested_file)
        items.append({
            "productCode": product["code"],
            "productDescription": product["description"],
            "supplier": product["supplier"],
            "brand": product["brand"],
            "factoryCode": product["factoryCode"],
            "photoUrl": product["photoUrl"],
            "photoDataUrl": product["photoDataUrl"],
            "suggestedFile": suggested_file,
            "score": best_score,
            "selected": selected,
        })

    return {
        "folderPath": str(folder),
        "productCount": len(products),
        "pdfCount": len(pdfs),
        "matchedCount": sum(1 for item in items if item["selected"]),
        "missingCount": sum(1 for item in items if not item["selected"]),
        "pdfOptions": [pdf["file"] for pdf in pdf_labels],
        "items": items,
    }


def rename_audited_product_pdfs(pdf_folder_path, renames):
    folder = Path(pdf_folder_path).expanduser()
    if not str(pdf_folder_path or "").strip() or not folder.is_dir():
        raise ValueError("Pasta de fichas invalida.")
    folder = folder.resolve()
    prepared = []
    seen_sources = set()
    seen_targets = set()

    for item in renames or []:
        source_name = str(item.get("sourceFile", "")).strip()
        code = normalize_code(item.get("productCode", ""))
        if not source_name or not code:
            continue
        source = (folder / source_name).resolve()
        if folder not in source.parents or not source.is_file() or source.suffix.lower() != ".pdf":
            raise ValueError(f"Ficha invalida ou nao encontrada: {source_name}")
        if source in seen_sources:
            raise ValueError(f"A mesma ficha foi selecionada mais de uma vez: {source_name}")

        clean_stem = re.sub(rf"^\s*{re.escape(code)}\s*[-_]\s*", "", source.stem).strip()
        safe_code = re.sub(r"[^A-Za-z0-9_-]+", "_", code).strip("_")
        if not safe_code:
            raise ValueError(f"Codigo de produto invalido para a ficha: {source_name}")
        target = source.with_name(f"{safe_code} - {clean_stem or 'Ficha'}.pdf")
        if target in seen_targets:
            raise ValueError(f"Duas fichas resultariam no mesmo nome: {target.name}")
        if target != source and target.exists():
            raise ValueError(f"Ja existe uma ficha chamada: {target.name}")
        prepared.append((source, target, source_name))
        seen_sources.add(source)
        seen_targets.add(target)

    renamed = []
    for source, target, source_name in prepared:
        if source != target:
            source.rename(target)
        renamed.append({
            "sourceFile": source_name,
            "targetFile": str(target.relative_to(folder)),
        })
    return {"renamedCount": len(renamed), "items": renamed}


def fill_workbook_with_ordered_pdfs(workbook_bytes, pdf_files):
    if not workbook_bytes:
        raise ValueError("Envie o Excel gerado para preencher.")
    if not any(pdf_files or []):
        raise ValueError("Envie pelo menos uma ficha PDF.")

    try:
        wb = load_workbook(io.BytesIO(workbook_bytes))
    except Exception as exc:
        raise ValueError("Nao consegui ler o Excel enviado.") from exc

    sheets = product_sheet_names(wb)
    if not sheets:
        raise ValueError("O Excel enviado nao possui abas de fichas para preencher.")

    products_ws = wb["Produtos"] if "Produtos" in wb.sheetnames else None
    brand_col = ensure_header_column(products_ws, "Marca") if products_ws else None
    ean_col = ensure_header_column(products_ws, "EAN") if products_ws else None
    dimensions_col = ensure_header_column(products_ws, "Dimensoes caixa") if products_ws else None

    for index, pdf_file in enumerate(pdf_files):
        if index >= len(sheets):
            break

        sheet = wb[sheets[index]]
        brand = detect_known_brand(
            sheet["B15"].value,
            sheets[index],
            pdf_file.get("name", "") if pdf_file else "",
        )
        pdf_data, _ = pdf_data_from_file(pdf_file)

        apply_pdf_data_to_product_sheet(sheet, pdf_data)
        if brand:
            sheet["B14"] = brand

        if products_ws:
            row = index + 2
            if brand and brand_col:
                products_ws.cell(row, brand_col, brand)
            if pdf_data.get("ean") and ean_col:
                products_ws.cell(row, ean_col, pdf_data["ean"])
            if pdf_data.get("box_dimensions") and dimensions_col:
                products_ws.cell(row, dimensions_col, pdf_data["box_dimensions"])

    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def normalize_match_text(value):
    value = unicodedata.normalize("NFKD", str(value or ""))
    value = value.encode("ascii", "ignore").decode("ascii")
    value = value.upper()
    value = re.sub(r"\bCONTRAFILE\b", "CONTRA FILE", value)
    value = re.sub(r"\bCOSTE\b", "COSTELA", value)
    value = re.sub(r"\bDESF\b", "DESFIADA", value)
    value = re.sub(r"\bHAMBURGUER\s+BOV\b", "HAMBURGUER CARNE BOVINA", value)
    value = re.sub(r"\bFLOW\b", "FLOWPACK", value)
    value = re.sub(r"\b(?:REV|V|VERSAO|REVISAO|PDF)\b\.?\s*\d*", " ", value)
    value = re.sub(r"[()]", " ", value)
    value = re.sub(r"[_\-]+", " ", value)
    value = re.sub(r"\bC\s*/\s*", " COM ", value)
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    tokens = [
        token
        for token in value.split()
        if len(token) > 1
        and (not re.fullmatch(r"\d+", token) or len(token) >= 5)
    ]
    return " ".join(tokens)


def supplier_pdf_label(path):
    path = Path(path)
    parts = [path.parent.name, path.stem]
    return normalize_match_text(" ".join(parts))


def supplier_pdf_text_label(path, max_chars=5000):
    path = Path(path)
    try:
        reader = PdfReader(str(path))
        text = "\n".join((page.extract_text() or "") for page in reader.pages[:2])
    except Exception:
        text = ""
    return normalize_match_text(" ".join([supplier_pdf_label(path), text[:max_chars]]))


def supplier_pdf_brand(path, label=""):
    path = Path(path)
    return detect_known_brand(str(path), label)


def supplier_pdf_internal_code(path):
    stem = Path(path).stem.strip()
    match = re.match(r"^(\d+)(?:\b|\s*[-_])", stem)
    return normalize_code(match.group(1)) if match else ""


def match_tokens(value):
    return {
        token for token in str(value or "").split()
        if len(token) > 2 and token not in MATCH_STOP_TOKENS
    }


def factory_code_is_strong(value):
    code = normalize_factory_code(value)
    if not code:
        return False
    if any(char.isalpha() for char in code):
        return len(code) >= 4
    return len(code) >= 5


def product_sheet_match_label(*values):
    text = normalize_match_text(" ".join(str(value or "") for value in values))
    tokens = text.split()
    # Keep repeated meaningful words out, but preserve order for SequenceMatcher.
    seen = set()
    out = []
    for token in tokens:
        if token in seen:
            continue
        seen.add(token)
        out.append(token)
    return " ".join(out)


def score_ordered_pdf_suggestion(
    product_label,
    pdf_label,
    product_brand="",
    pdf_brand="",
    product_factory_code="",
    product_internal_code="",
    pdf_internal_code="",
):
    if normalize_code(product_internal_code) and (
        normalize_code(product_internal_code) == normalize_code(pdf_internal_code)
    ):
        return 100

    product_tokens = match_tokens(product_label)
    pdf_tokens = match_tokens(pdf_label)
    if not product_tokens or not pdf_tokens:
        return 0

    product_brand_key = normalize_key(product_brand)
    pdf_brand_key = normalize_key(pdf_brand)
    if product_brand_key and pdf_brand_key and product_brand_key != pdf_brand_key:
        return 0

    factory_code = normalize_factory_code(product_factory_code)
    if factory_code and factory_code in pdf_tokens and factory_code_is_strong(factory_code):
        return 100 if product_brand_key and pdf_brand_key == product_brand_key else 96

    required_identity = product_tokens & PRODUCT_IDENTITY_TOKENS
    if required_identity:
        identity_overlap = required_identity & pdf_tokens
        if len(identity_overlap) < len(required_identity):
            return 0

    overlap = product_tokens & pdf_tokens
    required_overlap = 1 if required_identity and len(product_tokens) <= 3 else 2
    if len(overlap) < required_overlap:
        return 0

    coverage = len(overlap) / len(product_tokens)
    reverse_coverage = len(overlap) / len(pdf_tokens)
    sequence = difflib.SequenceMatcher(None, product_label, pdf_label).ratio()
    identity_coverage = (
        len(required_identity & pdf_tokens) / len(required_identity)
        if required_identity else 0
    )
    score = int((coverage * 60 + reverse_coverage * 15 + sequence * 10 + identity_coverage * 15))

    if product_brand_key and pdf_brand_key == product_brand_key:
        score += 12

    variant_tokens = {"ARG", "ARGENTINA", "FIT", "KIDS", "WAGYU", "MEDALHAO", "DIVIDIDO", "QUAD"}
    extra_variants = (pdf_tokens - product_tokens) & variant_tokens
    if extra_variants:
        score -= 35 * len(extra_variants)

    return max(0, min(100, score))


def prepare_supplier_match_products(rows):
    products = []
    for row in rows:
        summary = row_summary(row)
        label = normalize_match_text(
            " ".join([
                summary["description"],
                summary["supplier"],
                summary["brand"],
                summary["category"],
            ])
        )
        if not label:
            continue
        products.append({
            "row": row,
            "code": summary["code"],
            "description": summary["description"],
            "label": label,
            "tokens": match_tokens(label),
        })
    return products


def score_supplier_pdf_match_label(pdf_label, product_label):
    if not pdf_label or not product_label:
        return 0
    return int(difflib.SequenceMatcher(None, pdf_label, product_label).ratio() * 100)


def score_supplier_pdf_match(pdf_label, row):
    summary = row_summary(row)
    product_label = normalize_match_text(
        " ".join([
            summary["description"],
            summary["supplier"],
            summary["brand"],
            summary["category"],
        ])
    )
    if not pdf_label or not product_label:
        return 0

    return score_supplier_pdf_match_label(pdf_label, product_label)


def supplier_match_candidates(products, pdf_tokens, limit=60):
    if not pdf_tokens:
        return products[:limit]

    scored = []
    for product in products:
        overlap = len(product["tokens"] & pdf_tokens)
        if overlap:
            scored.append((overlap, product))

    strong = [item for item in scored if item[0] >= 2]
    usable = strong or scored
    usable.sort(key=lambda item: item[0], reverse=True)
    return [product for _, product in usable[:limit]]


def analyze_supplier_pdf_folder(pdf_folder_path, rows, min_score=55):
    folder = Path(pdf_folder_path).expanduser()
    if not str(pdf_folder_path or "").strip():
        return {
            "folderPath": "",
            "pdfCount": 0,
            "readOk": 0,
            "withEan": 0,
            "withDimensions": 0,
            "matched": 0,
            "unmatched": 0,
            "items": [],
            "errors": [],
        }
    if not folder.exists():
        raise ValueError(f"Pasta de fichas nao encontrada: {folder}")
    if not folder.is_dir():
        raise ValueError(f"O caminho das fichas nao e uma pasta: {folder}")

    pdfs = sorted(folder.rglob("*.pdf"))
    products = prepare_supplier_match_products(rows)
    items = []
    errors = []
    read_ok = 0
    with_ean = 0
    with_dimensions = 0

    for pdf in pdfs:
        rel_name = str(pdf.relative_to(folder))
        try:
            data = extract_supplier_pdf_data(pdf.read_bytes())
            read_ok += 1
            if data.get("ean"):
                with_ean += 1
            if data.get("box_dimensions"):
                with_dimensions += 1

            pdf_label = supplier_pdf_label(pdf)
            pdf_tokens = match_tokens(pdf_label)
            candidates = supplier_match_candidates(products, pdf_tokens)

            best_product = None
            best_score = 0
            for product in candidates:
                score = score_supplier_pdf_match_label(pdf_label, product["label"])
                if score > best_score:
                    best_score = score
                    best_product = product

            matched = best_product is not None and best_score >= min_score
            items.append({
                "file": rel_name,
                "ean": data.get("ean", ""),
                "boxDimensions": data.get("box_dimensions", ""),
                "altura": data.get("altura", ""),
                "largura": data.get("largura", ""),
                "comprimento": data.get("comprimento", ""),
                "matched": matched,
                "score": best_score,
                "productCode": best_product["code"] if matched else "",
                "productDescription": best_product["description"] if matched else "",
            })
        except Exception as exc:
            errors.append({"file": rel_name, "error": str(exc)})

    matched_count = sum(1 for item in items if item["matched"])
    return {
        "folderPath": str(folder),
        "pdfCount": len(pdfs),
        "readOk": read_ok,
        "withEan": with_ean,
        "withDimensions": with_dimensions,
        "matched": matched_count,
        "unmatched": len(items) - matched_count,
        "items": items[:200],
        "errors": errors[:50],
    }


def supplier_pdf_data_by_product(pdf_folder_path, rows, min_score=100):
    if not str(pdf_folder_path or "").strip():
        return {}

    folder = Path(pdf_folder_path).expanduser()
    if not folder.exists():
        raise ValueError(f"Pasta de fichas nao encontrada: {folder}")
    if not folder.is_dir():
        raise ValueError(f"O caminho das fichas nao e uma pasta: {folder}")

    best_by_code = {}
    products = []
    for row in rows:
        summary = row_summary(row)
        product_label = product_sheet_match_label(
            summary["description"],
            summary["supplier"],
            summary["brand"],
            summary["category"],
        )
        if product_label and summary["code"]:
            products.append({
                "code": summary["code"],
                "label": product_label,
                "brand": detect_known_brand(
                    summary["description"], summary["supplier"], summary["brand"]
                ),
                "factoryCode": summary["factoryCode"],
            })

    pdf_labels = []
    for pdf in sorted(folder.rglob("*.pdf")):
        label = supplier_pdf_text_label(pdf)
        pdf_labels.append({
            "path": pdf,
            "file": str(pdf.relative_to(folder)),
            "label": label,
            "brand": supplier_pdf_brand(pdf, label),
            "internalCode": supplier_pdf_internal_code(pdf),
        })

    for product in products:
        for pdf in pdf_labels:
            score = score_ordered_pdf_suggestion(
                product["label"],
                pdf["label"],
                product["brand"],
                pdf["brand"],
                product["factoryCode"],
                product["code"],
                pdf["internalCode"],
            )
            if score < min_score:
                continue
            current = best_by_code.get(product["code"])
            if current and current["score"] >= score:
                continue
            try:
                data = extract_supplier_pdf_data(pdf["path"].read_bytes())
            except ValueError:
                continue
            best_by_code[product["code"]] = {
                "score": score,
                "file": pdf["file"],
                "data": data,
            }

    return best_by_code


def image_code_from_name(name):
    base = os.path.basename(name)
    m = re.match(r"\s*(\d+)", base)
    return normalize_code(m.group(1)) if m else ""


def preferred_image_name(code, names):
    normalized = normalize_code(code)

    def sort_key(name):
        stem = Path(os.path.basename(name)).stem.strip()
        stem_without_copy = re.sub(r"\s*\(\d+\)\s*$", "", stem).strip()
        has_copy_suffix = stem_without_copy != stem
        exact_code_name = normalize_code(stem) == normalized

        if exact_code_name:
            priority = 0
        elif not has_copy_suffix and normalize_code(stem_without_copy) == normalized:
            priority = 1
        else:
            priority = 2

        return priority, has_copy_suffix, str(name).lower()

    return sorted(names, key=sort_key)[0]


def index_zip_images(zip_bytes):
    mapping = {}
    zf = None
    if not zip_bytes:
        return zf, mapping

    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile as exc:
        raise ValueError("O arquivo de fotos nao e um ZIP valido.") from exc

    for name in zf.namelist():
        if name.endswith("/"):
            continue

        ext = os.path.splitext(name)[1].lower()
        if ext not in IMAGE_EXTENSIONS:
            continue

        code = image_code_from_name(name)
        if code:
            mapping.setdefault(code, []).append(name)

    return zf, mapping


def index_folder_images(folder: str):
    folder = Path(folder).expanduser()
    mapping = {}

    if not folder.exists():
        raise ValueError(f"Pasta de fotos nao encontrada: {folder}")
    if not folder.is_dir():
        raise ValueError(f"O caminho informado nao e uma pasta: {folder}")

    for p in folder.rglob("*"):
        if p.suffix.lower() in IMAGE_EXTENSIONS:
            code = image_code_from_name(p.name)
            if code:
                mapping.setdefault(code, []).append(str(p))

    return mapping


def _folder_cache_key(folder: str):
    folder_path = Path(folder).expanduser().resolve()
    stat = folder_path.stat()
    return str(folder_path), stat.st_mtime_ns


@lru_cache(maxsize=8)
def _cached_index_folder_images(folder: str, mtime_ns: int):
    return index_folder_images(folder)


def index_folder_images_cached(folder: str):
    folder_path = Path(folder).expanduser()
    if not folder_path.exists():
        raise ValueError(f"Pasta de fotos nao encontrada: {folder_path}")
    if not folder_path.is_dir():
        raise ValueError(f"O caminho informado nao e uma pasta: {folder_path}")

    folder, mtime_ns = _folder_cache_key(str(folder_path))
    return _cached_index_folder_images(folder, mtime_ns)


def default_folder_path(folder_path=""):
    if folder_path and str(folder_path).strip():
        return str(folder_path).strip()
    if DEFAULT_PHOTOS.exists():
        return str(DEFAULT_PHOTOS)
    return ""


def prepare_image_file(code, zip_file, zip_mapping, folder_mapping, temp_dir):
    code = normalize_code(code)

    if code in folder_mapping:
        return preferred_image_name(code, folder_mapping[code])

    if zip_file and code in zip_mapping:
        name = preferred_image_name(code, zip_mapping[code])
        ext = os.path.splitext(name)[1] or ".jpg"
        out = Path(temp_dir) / f"{code}{ext}"
        out.write_bytes(zip_file.read(name))
        return str(out)

    return None


def optimize_image_for_excel(path, out_path, max_width, max_height, quality=72):
    try:
        with PILImage.open(path) as source:
            img = ImageOps.exif_transpose(source)
            img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)

            if img.mode in ("RGBA", "LA") or (
                img.mode == "P" and "transparency" in img.info
            ):
                background = PILImage.new("RGB", img.size, "white")
                background.paste(img, mask=img.convert("RGBA").split()[-1])
                img = background
            else:
                img = img.convert("RGB")

            img.save(out_path, format="JPEG", quality=quality, optimize=True)
            return str(out_path), img.size
    except (OSError, UnidentifiedImageError):
        return None, None


def build_excel_images(code, img_path, temp_dir):
    if not img_path:
        return {}

    safe_code = re.sub(r"[^A-Za-z0-9_-]+", "_", code or Path(img_path).stem)
    image_dir = Path(temp_dir) / "excel_images"
    image_dir.mkdir(exist_ok=True)

    images = {}
    for name, dimensions in {
        "summary": (150, 105),
        "sheet": (270, 270),
    }.items():
        out_path = image_dir / f"{safe_code}_{name}.jpg"
        optimized_path, size = optimize_image_for_excel(
            img_path,
            out_path,
            dimensions[0],
            dimensions[1],
        )
        if optimized_path:
            images[name] = {
                "path": optimized_path,
                "width": size[0],
                "height": size[1],
            }

    return images


def image_from_info(image_info, fallback_width, fallback_height):
    img = XLImage(image_info["path"])
    img.width = image_info.get("width") or fallback_width
    img.height = image_info.get("height") or fallback_height
    return img


def excel_column_width_to_pixels(width):
    width = 8.43 if width is None else float(width)
    if width < 1:
        return int(width * 12)
    return int(width * 7 + 5)


def excel_range_size_pixels(ws, range_ref):
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)

    width = 0
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        width += excel_column_width_to_pixels(ws.column_dimensions[letter].width)

    height = 0
    default_height = ws.sheet_format.defaultRowHeight or 15
    for row in range(min_row, max_row + 1):
        row_height = ws.row_dimensions[row].height or default_height
        height += points_to_pixels(row_height)

    return width, height


def add_centered_image(ws, image_info, range_ref, fallback_width, fallback_height):
    img = image_from_info(image_info, fallback_width, fallback_height)
    min_col, min_row, _, _ = range_boundaries(range_ref)
    box_width, box_height = excel_range_size_pixels(ws, range_ref)

    offset_x = max(0, int((box_width - img.width) / 2))
    offset_y = max(0, int((box_height - img.height) / 2))

    img.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=min_col - 1,
            row=min_row - 1,
            colOff=pixels_to_EMU(offset_x),
            rowOff=pixels_to_EMU(offset_y),
        ),
        ext=XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height)),
    )
    ws.add_image(img)


def safe_sheet_name(name):
    name = re.sub(r"[\\/*?:\[\]]", "-", str(name))[:31]
    return name or "Produto"


def unique_sheet_name(wb, name):
    base = safe_sheet_name(name)
    if base not in wb.sheetnames:
        return base

    counter = 2
    while True:
        suffix = f" ({counter})"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in wb.sheetnames:
            return candidate
        counter += 1


def val(row, *names):
    for n in names:
        if n in row and str(row[n]).strip():
            return str(row[n]).strip()

    normalized_names = {normalize_key(n) for n in names}
    for key, value in row.items():
        if normalize_key(key) in normalized_names and str(value).strip():
            return str(value).strip()

    return ""


def row_code(row):
    return normalize_code(val(row, "Codigo", "Codigo interno"))


def row_summary(row, image_codes=None, photo_data_url="", photo_version=""):
    image_codes = image_codes or set()
    code = row_code(row)
    photo_url = ""
    if code in image_codes and not photo_data_url:
        photo_url = f"/api/photo/{code}"
        if photo_version:
            photo_url = f"{photo_url}?v={photo_version}"

    return {
        "code": code,
        "description": val(
            row,
            "Descricao",
            "Descricao do produto",
            "Nome do produto",
            "Nome Produto",
            "Produto",
        ),
        "supplier": val(row, "Nome do fornecedor", "Fornecedor", "Razao Social"),
        "brand": val(row, "Marca"),
        "factoryCode": normalize_factory_code(val(row, "Codigo fabrica", "CODIGO_FABRICA")),
        "category": val(row, "Nome da categoria"),
        "package": val(row, "Embalagem"),
        "ncm": val(row, "NCM", "NCM + Excecao"),
        "ean": val(row, "GTIN Unid.Venda", "EAN Unid. Tributavel"),
        "boxDimensions": val(row, "Dimensoes da caixa", "Dimensões da caixa"),
        "hasPhoto": code in image_codes,
        "photoUrl": photo_url,
        "photoDataUrl": photo_data_url,
    }


def build_search_text(row):
    summary = row_summary(row)
    return " ".join([
        summary["code"],
        summary["description"],
        summary["supplier"],
        summary["brand"],
    ]).lower()

def read_product_prices(region=DEFAULT_PRICE_REGION):
    if not PRICE_FILE.exists():
        raise ValueError(f"Planilha de preços não encontrada: {PRICE_FILE.name}")

    wb = load_workbook(PRICE_FILE, read_only=True, data_only=True)
    ws = wb["preço"] if "preço" in wb.sheetnames else wb.active

    headers = {
        normalize_key(cell.value): index
        for index, cell in enumerate(next(ws.iter_rows()))
    }

    code_index = headers["codprod"]
    region_index = headers["numregiao"]
    price_index = headers["precovenda"]

    prices = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = normalize_code(row[code_index])
        row_region = row[region_index]
        price = row[price_index]

        if code and str(row_region) == str(region) and price is not None:
            prices[code] = price

    wb.close()
    return prices

def preencher_ficha_template(ficha, row, code, sheet_image, price=None):
    descricao = row_summary(row)["description"]
    marca = val(row, "Marca")
    embalagem = val(row, "Embalagem")
    ncm = val(row, "NCM", "NCM + Excecao")
    cest = val(row, "CEST")

    ean = (
        val(row, "GTIN Unid.Venda")
        or val(row, "EAN Unid. Tributavel")
        or val(row, "GTIN Unid. Tributavel")
    )

    peso_liq = val(row, "Peso liq.")
    peso_bruto = val(row, "Peso bruto")
    validade = val(row, "Dias Validade")
    box_dimensions = val(row, "Dimensoes da caixa", "Dimensões da caixa")
    parsed_dimensions = parse_box_dimensions(box_dimensions)
    altura = val(row, "Unid.Altura(cm)", "Altura(cm)", "Altura") or parsed_dimensions.get("altura", "")
    largura = val(row, "Unid.Largura(cm)", "Largura(cm)", "Largura") or parsed_dimensions.get("largura", "")
    comprimento = (
        val(row, "Unid.Comprim(cm)", "Comprimento(cm)", "Comprimento")
        or parsed_dimensions.get("comprimento", "")
    )

    ficha["B14"] = marca
    ficha["B15"] = descricao
    ficha["B16"] = price if price is not None else ""
    ficha["B16"].number_format = 'R$ #,##0.00'
    ficha["B18"] = ncm
    ficha["B19"] = cest
    ficha["B20"] = "0,00"
    ficha["B21"] = "0"
    ficha["B22"] = "Nacional"
    ficha["B23"] = "CIF"
    ficha["B24"] = embalagem
    ficha["B26"] = altura
    ficha["B27"] = largura
    ficha["B28"] = comprimento
    ficha["B29"] = peso_liq
    ficha["B30"] = peso_bruto
    ficha["D29"] = validade
    ficha["A32"] = descricao
    ficha["D32"] = code
    ficha["F32"] = ean
    ficha["B46"] = ncm
    ficha._images = []

    if sheet_image:
        try:
            add_centered_image(ficha, sheet_image, "F9:H21", 270, 270)
        except Exception:
            pass


def load_inputs(csv_bytes=None, zip_bytes=None, folder_path=""):
    if csv_bytes:
        _, rows = read_csv_bytes(csv_bytes)
    elif DEFAULT_CSV.exists():
        _, rows = read_csv_bytes(DEFAULT_CSV.read_bytes())
    elif find_template_file().exists():
        _, rows = read_xlsx_products(find_template_file())
    else:
        raise ValueError("Envie o CSV ou coloque uma planilha .xlsx de produtos na pasta do sistema.")

    rows = apply_factory_codes(rows)

    zip_file, zip_mapping = index_zip_images(zip_bytes)
    if not zip_bytes and DEFAULT_ZIP.exists():
        zip_file, zip_mapping = index_zip_images(DEFAULT_ZIP.read_bytes())

    folder_mapping = {}
    resolved_folder = default_folder_path(folder_path)
    if resolved_folder:
        folder_mapping = index_folder_images_cached(resolved_folder)

    return rows, zip_file, zip_mapping, folder_mapping


def find_photo_for_code(code, folder_path=""):
    mapping = {}
    resolved_folder = default_folder_path(folder_path)
    if resolved_folder:
        mapping = index_folder_images_cached(resolved_folder)

    normalized = normalize_code(code)
    if normalized in mapping:
        return preferred_image_name(normalized, mapping[normalized])

    return None


def image_bytes_to_data_url(data, max_size=(260, 260), quality=72):
    try:
        with PILImage.open(io.BytesIO(data)) as source:
            img = ImageOps.exif_transpose(source)
            img.thumbnail(max_size, PILImage.Resampling.LANCZOS)

            if img.mode in ("RGBA", "LA") or (
                img.mode == "P" and "transparency" in img.info
            ):
                background = PILImage.new("RGB", img.size, "white")
                background.paste(img, mask=img.convert("RGBA").split()[-1])
                img = background
            else:
                img = img.convert("RGB")

            out = io.BytesIO()
            img.save(out, format="JPEG", quality=quality, optimize=True)
    except (OSError, UnidentifiedImageError):
        return ""

    encoded = base64.b64encode(out.getvalue()).decode("ascii")
    return f"data:image/jpeg;base64,{encoded}"


def zip_preview_data_url(code, zip_file, zip_mapping):
    if not zip_file or code not in zip_mapping:
        return ""

    name = preferred_image_name(code, zip_mapping[code])
    try:
        return image_bytes_to_data_url(zip_file.read(name))
    except (KeyError, RuntimeError, zipfile.BadZipFile):
        return ""


def photo_file_version(path):
    try:
        stat = Path(path).stat()
    except OSError:
        return ""
    return f"{stat.st_mtime_ns}-{stat.st_size}"


def list_products(
    csv_bytes=None,
    zip_bytes=None,
    folder_path="",
    search="",
    only_with_photo=False,
    page=1,
    page_size=120,
):
    rows, zip_file, zip_mapping, folder_mapping = load_inputs(csv_bytes, zip_bytes, folder_path)
    image_codes = set(zip_mapping) | set(folder_mapping)
    filtered_rows = []
    query = search.lower().strip()
    page = max(1, int(page or 1))
    page_size = min(500, max(1, int(page_size or 120)))

    for row in rows:
        if query and query not in build_search_text(row):
            continue
        code = row_code(row)
        item = row_summary(row, image_codes)
        if only_with_photo and not item["hasPhoto"]:
            continue
        filtered_rows.append(row)

    start = (page - 1) * page_size
    end = start + page_size
    products = []

    prices = read_product_prices() if PRICE_FILE.exists() else {}

    for row in filtered_rows[start:end]:
        code = row_code(row)
        photo_data_url = ""
        photo_version = ""
        if code in folder_mapping:
            photo_version = photo_file_version(preferred_image_name(code, folder_mapping[code]))
        if code not in folder_mapping:
            photo_data_url = zip_preview_data_url(code, zip_file, zip_mapping)
        item = row_summary(row, image_codes, photo_data_url, photo_version)
        item["originalPrice"] = prices.get(code)
        products.append(item)

    return {
        "total": len(rows),
        "filteredTotal": len(filtered_rows),
        "page": page,
        "pageSize": page_size,
        "hasNextPage": end < len(filtered_rows),
        "photoCount": len(image_codes),
        "templateFound": find_template_file().exists(),
        "templateName": find_template_file().name if find_template_file().exists() else "",
        "products": products,
    }


def create_workbook_bytes(
    selected_rows,
    zip_file,
    zip_mapping,
    folder_mapping,
    include_product_sheets=True,
    price_by_code=None,
):
    with TemporaryDirectory(prefix="produto_fotos_") as temp_dir:
        return _create_workbook_bytes(
            selected_rows,
            zip_file,
            zip_mapping,
            folder_mapping,
            temp_dir,
            include_product_sheets,
            price_by_code=price_by_code,
        )


def _create_workbook_bytes(
    selected_rows,
    zip_file,
    zip_mapping,
    folder_mapping,
    temp_dir,
    include_product_sheets=True,
    price_by_code=None,
):
    price_by_code = price_by_code or {}
    template_file = find_template_file()
    if template_file.exists():
        wb = load_workbook(template_file)
    else:
        wb = Workbook()

    modelo = wb["base"] if "base" in wb.sheetnames else wb[wb.sheetnames[0]]

    if "Produtos" in wb.sheetnames:
        del wb["Produtos"]

    ws = wb.create_sheet("Produtos", 0)

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    headers = [
        "Imagem",
        "Codigo",
        "Descricao",
        "Fornecedor",
        "Marca",
        "Categoria",
        "Embalagem",
        "NCM",
        "EAN",
        "Dimensoes caixa",
        "Foto encontrada?",
    ]
    ws.append(headers)

    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")

    widths = [24, 12, 55, 36, 20, 25, 18, 14, 20, 24, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    for idx, row in enumerate(selected_rows, start=2):
        summary = row_summary(row)
        code = summary["code"]

        img_path = None
        if code:
            img_path = prepare_image_file(
                code,
                zip_file,
                zip_mapping,
                folder_mapping,
                temp_dir,
            )
        excel_images = build_excel_images(code, img_path, temp_dir)

        ws.row_dimensions[idx].height = 95
        values = [
            "",
            code,
            summary["description"],
            summary["supplier"],
            summary["brand"],
            summary["category"],
            summary["package"],
            summary["ncm"],
            summary["ean"],
            summary["boxDimensions"],
            "SIM" if excel_images else "NAO",
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(idx, col, value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        if excel_images.get("summary"):
            try:
                ws.add_image(image_from_info(excel_images["summary"], 150, 105), f"A{idx}")
            except Exception:
                ws.cell(idx, 10, "ERRO NA FOTO")

        if not include_product_sheets:
            continue

        nome_aba = unique_sheet_name(
            wb,
            f"{code or 'Produto'} - {summary['description'][:24]}",
        )
        ficha = wb.copy_worksheet(modelo)
        ficha.title = nome_aba
        preencher_ficha_template(
            ficha,
            row,
            code,
            excel_images.get("sheet"),
            price_by_code.get(code),
        )

    if modelo.title in wb.sheetnames:
        del wb[modelo.title]

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_workbook_for_codes(
    selected_codes,
    csv_bytes=None,
    zip_bytes=None,
    folder_path="",
    include_product_sheets=True,
    supplier_pdf_bytes=None,
    supplier_pdf_folder_path="",
    include_prices=False,
    price_overrides=None,
):
    rows, zip_file, zip_mapping, folder_mapping = load_inputs(csv_bytes, zip_bytes, folder_path)
    selected = set(normalize_code(code) for code in selected_codes)
    selected_rows = [row for row in rows if row_code(row) in selected]
    if not selected_rows:
        raise ValueError("Nenhum produto selecionado.")
    if supplier_pdf_bytes:
        if len(selected_rows) != 1:
            raise ValueError("Para usar ficha PDF do fornecedor, selecione apenas um produto.")
        selected_rows = [apply_supplier_pdf_data(selected_rows[0], extract_supplier_pdf_data(supplier_pdf_bytes))]
    elif supplier_pdf_folder_path:
        mapped_pdf_data = supplier_pdf_data_by_product(supplier_pdf_folder_path, selected_rows)
        selected_rows = [
            apply_supplier_pdf_data(row, mapped_pdf_data[row_code(row)]["data"])
            if row_code(row) in mapped_pdf_data
            else row
            for row in selected_rows
        ]
    price_by_code = read_product_prices() if include_prices else {}
    price_by_code.update(price_overrides or {})

    return create_workbook_bytes(
        selected_rows,
        zip_file,
        zip_mapping,
        folder_mapping,
        include_product_sheets,
        price_by_code=price_by_code,
    )
